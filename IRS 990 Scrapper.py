from bs4 import BeautifulSoup
import urllib
import requests
import pandas as pd
import locale
import openpyxl
from pprint import pprint 
import json 

# Set locale to the US for currency formatting
locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')

# To-Do List
# Finish Commenting Functions
# Better Way of searching for colleges
#   Probably going to be some form of web creation at this point
# Fix Random College Bug
#   Some colleges have no tax fillings for the "college" but they do for things like "Alumni Association Inc"
#   Add break in code that returns (Could not find) if tax filings contain keywords like above
# Question
#   Code currently removes anyone with no comp (0 Base Comp, 0 Other Comp)
#   The "Number of employees" value is based on ALL employees (Including those with no comp)
#   Should I
#     A) Report all reported employees regardless of comp (Bloat the file)
#     B) Not count those 0 comp reported people in the "total employee count"

# Use the information from ProPublica API and more GET requests to get the xml version of the IRS 990 form
def get_irs_990_web_content(search_result):

    # Get the identification number for the institution to be used in another url
    search_result_json = search_result.json()
    search_result_ein = search_result_json['organizations'][0]['ein']

    # Search for the base html page of the searched Non-profit 
    specific_nonprofit_search = "https://projects.propublica.org/nonprofits/organizations/" + str(search_result_ein)
    page = requests.get(specific_nonprofit_search).content

    # Use BeautifulSoup library to find the first xml button on the website and get the link to the xml report
    page_soup = BeautifulSoup(page, "html.parser")
    xml_soup = page_soup.find_all(lambda tag: tag.name == "a" and "XML" in tag.text)
    
    all_xml_found = []
    stopper = 0
    for href in xml_soup:
        if stopper == 10:
            break
        link = href.get("href")
        if '/nonprofits/download-xml' in link:
            all_xml_found.append("https://projects.propublica.org" + link)
        stopper += 1

    return all_xml_found

# Get all the information present on the IRS990 form on the occupations paid the greatest in the college
def get_institution_occupation_data(irs_990_soup):
    # We will be looking for these specific pieces of information
    # "Title Group" is the generalization of "Title" determined in a function below
    # "Total Compensation" is the addition of base comp and other comp
    columns = ["Year", "Name", "Title", "Title Group", "Base Compensation", "Other Comp", "Total Comp"]
    jobs = pd.DataFrame(columns=columns)
    company_wide_compensation = 0
    total_reported_employees = 0
    
    # Get the year of the data
    endYear = irs_990_soup.find("TaxPeriodEndDt").text.strip()[:4]

    # Get the IRS990 part of the xml tax filling
    tax_filing_soup = irs_990_soup.find("IRS990")
    # PartVII in the tax filling is where all the occupational information is
    occupation_soup = tax_filing_soup.find_all("Form990PartVIISectionAGrp")
    for employee_soup in occupation_soup:
        name = ""
        # Forms have the name of an employee in different tasks, so try both
        try:
            name = employee_soup.find("PersonNm").text.strip()
        except:
            business_soup = employee_soup.find("BusinessName")
            name = business_soup.find("BusinessNameLine1Txt").text.strip()
        
        try:
            base_comp = int(employee_soup.find("ReportableCompFromOrgAmt").text.strip())
        except:
            base_comp = 0
        try:
            other_comp = int(employee_soup.find("OtherCompensationAmt").text.strip())
        except:
            other_comp = 0
        total_comp = base_comp + other_comp
        # Store the total compensation of every employee listed in the tax form
        company_wide_compensation += total_comp
        
        # If the employee is listed and doesn't make any money, skip them
        if total_comp == 0:
            continue

        total_reported_employees += 1

        title_name = employee_soup.find("TitleTxt").text.strip()
        # Use the title to get a generalization of the name 
        title_group = get_title_group(title_name)

        job_info = {
            "Year" : endYear,
            "Name" : name,
            "Title": title_name,
            "Title Group" : title_group,
            "Base Compensation" : base_comp,
            "Other Comp" : other_comp,
            "Total Comp" : total_comp
        }
        # Add the employee's information to the list of jobs in the college
        jobs.loc[len(jobs)] = job_info

    # Sort all the employees by total compensation descending
    top_jobs = jobs.sort_values(by=['Total Comp'], ascending=False)
    return top_jobs.to_dict(), company_wide_compensation, total_reported_employees

# Use the title of an employee to generalize it into a group that can be shared by multiple
# "PRESIDENT (THROUGH 8/12)" and "PRESIDENT/CEO" get simplified into "President"
def get_title_group(title_name):
    # Name the base title group as other, so that if no title is found, report it as other
    tg = "Other"
    t = title_name.lower()
    # As of now, the code will look through this list going down. If it finds one title, it will skip the rest
    # If a title is "SECRETARY/VP/TREAS", it will be reported as "VP"
    list_of_titles = [
        "Vice President",
        "Vice Provost",
        "President", 
        "Provost", 
        "VP", 
        "Trustee",
        "Dean", 
        "Exec", 
        "Prof",
        "Treas",
        "Secretary",
        "Chief",
        "Dept Head"
    ]
    
    for title in list_of_titles:
        if title.lower() in t:
            tg = title
            break
    return tg

# Get the overarching revenue data of a college
def get_institution_revenue_data(irs_990_soup, nonprofit_query):
    endYear = irs_990_soup.find("TaxPeriodEndDt").text.strip()[:4]

    filer = irs_990_soup.find("Filer")
    bName = filer.find("BusinessName")
    try:
        company = bName.find("BusinessNameLine1Txt").text.strip()
    except:
        company = nonprofit_query

    pyTotalRevenue = int(irs_990_soup.find("PYTotalRevenueAmt").text.strip())
    cyTotalRevenue = int(irs_990_soup.find("CYTotalRevenueAmt").text.strip())

    pyNetRevenue = int(irs_990_soup.find("PYRevenuesLessExpensesAmt").text.strip())
    cyNetRevenue = int(irs_990_soup.find("CYRevenuesLessExpensesAmt").text.strip())

    revenue_dict = {
        "year" : endYear,
        "company" : company,
        "pyTotalRevenue" : pyTotalRevenue,
        "cyTotalRevenue" : cyTotalRevenue,
        "pyNetRevenue" : pyNetRevenue,
        "cyNetRevenue" : cyNetRevenue,
        "company_wide_compensation" : 0,
        "average_comp_per_reported" : 0,
        "net_over_comp_index" : 0,
        "total_reported_employees" : 0,
        "job_information" : pd.DataFrame()
    }

    return revenue_dict

# Write all the scrapped information to an excel file
def write_intitution_to_excel(revenue_dict, nonprofit_subtitle, filename, job_list_index):
    
    # If a college or list of college excel file already exists for the list or single college, update the list instead of creating a new one
    try:
        excelWorkbook = openpyxl.load_workbook(filename)
    except:
        excelWorkbook = openpyxl.Workbook()

    # If a college in the excel file already exists, overwrite the sheet instead of making a new one
    try:
        sheet = excelWorkbook[nonprofit_subtitle]
    except:
        excelWorkbook.create_sheet(title=nonprofit_subtitle)
        sheet = excelWorkbook[nonprofit_subtitle]

    sheet = excelWorkbook["Sheet"]
    sheet["A1"] = "Year"
    sheet["B1"] = "School"
    sheet["C1"] = "Revenue"
    sheet["D1"] = "Net Income"
    sheet["E1"] = "Total Employee Comp"
    sheet["F1"] = "Average Comp"
    sheet["G1"] = "Net/Comp Index"
    sheet["H1"] = "Total Employees"
    sheet["I1"] = "Presidents"
    sheet["J1"] = "Vice Presidents"
    sheet["K1"] = "Provosts"
    sheet["L1"] = "Trustees"
    sheet["M1"] = "Deans"
    sheet["N1"] = "Executives"
    sheet["O1"] = "Professors"
    sheet["P1"] = "Treasurers"
    sheet["Q1"] = "Secretaries"
    sheet["R1"] = "Chiefs"
    sheet["S1"] = "Dept Heads"
    sheet["T1"] = "Other"

    sheet = excelWorkbook[nonprofit_subtitle]
    sheet["A1"] = "Year Range"
    sheet["A2"] = "Current Year Total Revenue"
    sheet["A3"] = "Prior Year Total Revenue"
    sheet["A4"] = "Current Year Net Income Less Expense"
    sheet["A5"] = "Prior Year Net Income Less Expense"
    sheet["A6"] = "Total College Wide Compensation"
    sheet["A7"] = "Average Comp Per Reported Employee"
    sheet["A8"] = "Net Income / Total Comp Index"
    sheet["A9"] = "Total Reported Employees"
    sheet["A11"] = "Year"
    sheet["B11"] = "Name"
    sheet["C11"] = "Title"
    sheet["D11"] = "Title Group"
    sheet["E11"] = "Base Compensation"
    sheet["F11"] = "Other Comp"
    sheet["G11"] = "Total Comp"

    summary_index = job_list_index
    college_summary_year_index = 66
    excel_row = 12
    for year_data in revenue_dict:
        sheet = excelWorkbook[nonprofit_subtitle]
        sheet[chr(college_summary_year_index) + "1"] = year_data["year"]
        sheet[chr(college_summary_year_index) + "2"] = locale.currency(year_data["cyTotalRevenue"], grouping=True)
        sheet[chr(college_summary_year_index) + "3"] = locale.currency(year_data["pyTotalRevenue"], grouping=True)
        sheet[chr(college_summary_year_index) + "4"] = locale.currency(year_data["cyNetRevenue"], grouping=True)
        sheet[chr(college_summary_year_index) + "5"] = locale.currency(year_data["pyNetRevenue"], grouping=True)
        sheet[chr(college_summary_year_index) + "6"] = locale.currency(year_data["company_wide_compensation"], grouping=True)
        sheet[chr(college_summary_year_index) + "7"] = locale.currency(year_data["average_comp_per_reported"], grouping=True)
        sheet[chr(college_summary_year_index) + "8"] = year_data["net_over_comp_index"]
        sheet[chr(college_summary_year_index) + "9"] = year_data["total_reported_employees"]
        
        college_summary_year_index += 1

        # Create a dictionary of the title groups to count how many times each group shows up in the school irs form
        num_of_titles = {
            "Vice President" : 0 ,
            "Vice Provost" : 0,
            "President" : 0, 
            "Provost" : 0, 
            "VP" : 0, 
            "Trustee" : 0,
            "Dean" : 0, 
            "Exec" : 0, 
            "Prof" : 0,
            "Treas" : 0,
            "Secretary" : 0,
            "Chief" : 0,
            "Dept Head" : 0,
            "Other" : 0
        }

        for job in range(0, len(year_data["job_information"]["Year"])):
            sheet["A"+str(excel_row)] = year_data["job_information"]["Year"][job]
            sheet["B"+str(excel_row)] = year_data["job_information"]["Name"][job]
            sheet["C"+str(excel_row)] = year_data["job_information"]["Title"][job]
            sheet["D"+str(excel_row)] = year_data["job_information"]["Title Group"][job]
            num_of_titles[year_data["job_information"]["Title Group"][job]] = num_of_titles.get(year_data["job_information"]["Title Group"][job]) + 1
            sheet["E"+str(excel_row)] = year_data["job_information"]["Base Compensation"][job]
            sheet["F"+str(excel_row)] = year_data["job_information"]["Other Comp"][job]
            sheet["G"+str(excel_row)] = year_data["job_information"]["Total Comp"][job]
            excel_row += 1

        # Add college summary information to the Master Sheet
        sheet = excelWorkbook["Sheet"]

        sheet["A" + str(summary_index)] = year_data["year"]
        sheet["B" + str(summary_index)] = nonprofit_subtitle
        sheet["C" + str(summary_index)] = locale.currency(year_data["cyTotalRevenue"], grouping=True)
        sheet["D" + str(summary_index)] = locale.currency(year_data["cyNetRevenue"], grouping=True)
        sheet["E" + str(summary_index)] = locale.currency(year_data["company_wide_compensation"], grouping=True)
        sheet["F" + str(summary_index)] = locale.currency(year_data["average_comp_per_reported"], grouping=True)
        sheet["G" + str(summary_index)] = year_data["net_over_comp_index"]
        sheet["H" + str(summary_index)] = year_data["total_reported_employees"]
        sheet["I" + str(summary_index)] = num_of_titles["President"]
        sheet["J" + str(summary_index)] = num_of_titles["Vice President"] + num_of_titles["VP"]
        sheet["K" + str(summary_index)] = num_of_titles["Provost"] + num_of_titles["Vice Provost"]
        sheet["L" + str(summary_index)] = num_of_titles["Trustee"]
        sheet["M" + str(summary_index)] = num_of_titles["Dean"]
        sheet["N" + str(summary_index)] = num_of_titles["Exec"]
        sheet["O" + str(summary_index)] = num_of_titles["Prof"]
        sheet["P" + str(summary_index)] = num_of_titles["Treas"]
        sheet["Q" + str(summary_index)] = num_of_titles["Secretary"]
        sheet["R" + str(summary_index)] = num_of_titles["Chief"]
        sheet["S" + str(summary_index)] = num_of_titles["Dept Head"]
        sheet["T" + str(summary_index)] = num_of_titles["Other"]
        
        summary_index += 1

    # Save the created workbook into the filename provided
    excelWorkbook.save(filename)

# "Main" function. Searches for a college or list of colleges
def search(nonprofit_search_query, nonprofit_subtitle, show_summary, filename, index):
    # Use ProPublica API to get basic institution data
    search_url = "https://projects.propublica.org/nonprofits/api/v2/search.json?"
    parameters = urllib.parse.urlencode({
        'q' : nonprofit_search_query
    })
    search_result = requests.get(search_url+parameters)

    # If an institution is found, continue; break if nothing is found
    if search_result.status_code == 200:
        print("Found", nonprofit_search_query)

        school_data_array = []

        # Get the web content in the form of an xml version of an IRS 990 form
        all_years_irs_990_soup = get_irs_990_web_content(search_result)
        
        for year_data in all_years_irs_990_soup:

            xml_search_result = requests.get(year_data).content

            # Use BeautifulSoup to parse the web content as xml and return
            irs_990_soup = BeautifulSoup(xml_search_result, features="xml")
            
            if (irs_990_soup.find("TaxPeriodEndDt") is None) and (irs_990_soup.find("IRS990").find("Form990PartVIISectionAGrp") is None):
                continue

            # Use xml content to get the paid reported occupations sorted by highest compensation descending
            top_jobs, company_wide_compensation, total_reported_employees = get_institution_occupation_data(irs_990_soup)

            # Get the basic IRS data from the xml content
            revenue_dict = get_institution_revenue_data(irs_990_soup, nonprofit_search_query)
            revenue_dict["company_wide_compensation"] = company_wide_compensation
            try:
                revenue_dict["average_comp_per_reported"] = company_wide_compensation / total_reported_employees
            except:
                revenue_dict["average_comp_per_reported"] = 0
            try:
                revenue_dict["net_over_comp_index"] = revenue_dict["cyNetRevenue"] / company_wide_compensation
            except:
                revenue_dict["net_over_comp_index"] = 0
            revenue_dict["total_reported_employees"] = total_reported_employees
            revenue_dict["job_information"] = top_jobs
            
            school_data_array.append(revenue_dict)

        # Write all the information to the Master Excel file
        write_intitution_to_excel(school_data_array, nonprofit_subtitle, filename, index)
        return len(school_data_array)
        
    else:
        # If the college is not found, report it
        print("Could not find", nonprofit_search_query)
        return 0


print("If reading from list, enter the filename. If not, just hit enter.")
using_list = input("Using list?: ")
# Make sure that if a list is used, that is is present in a .txt file format
# See "Institutes.txt" for an example
if len(using_list) > 4 and using_list[-4:] == ".txt":
    print("Reading from file", using_list, "...")
    institutes = []
    # Get the list of institutes and institute subtitles
    with open("Inputs/"+using_list, "r") as f:
        institutes = f.readlines()
    # Every line has a "/n" indicating a new line in the text file, so get rid of them
    for i in range(len(institutes)):
        institutes[i] = institutes[i][:-1]
    
    # Look through every even numbered line for the full institute name to be searched (Odd is the subtitle)
    summary_page_index = 2
    for i in range(len(institutes)):
        if i % 2 == 0:
            # Every college will be saved in a sheet in a master excel file, named after the name of the txt file the list is in
            # Ex. Using the lists of colleges in "Institutes.txt" will output in "Institutes.xlsx"
            summary_page_index += search(institutes[i], institutes[i+1], False, "Outputs/"+using_list[:-4]+".xlsx", summary_page_index)
        else:
            continue
    print("Output college information in", "Outputs/"+using_list[:-4]+".xlsx")

else:
    # Ask for Institution Name + Institution Subtitle (Excel Sheet Name)
    nonprofit_search_query = input("Full Institute Name: ")
    nonprofit_subtitle = input("Institute Nickname: ")
    # If no subtitle is provided, just make it the search query
    if nonprofit_subtitle == "":
        nonprofit_subtitle = nonprofit_search_query
    # The output file for a single query will be stored in the input subtitles name + .xlsx
    #Ex. If I look up Worcester Polytechnic Institute (WPI), it will be saved as WPI.xlsx
    search(nonprofit_search_query, nonprofit_subtitle, False, "Outputs/"+nonprofit_subtitle+".xlsx",2)
    print("Output college information in", "Outputs/"+nonprofit_subtitle+".xlsx")

